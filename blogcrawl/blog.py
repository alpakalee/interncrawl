from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver import ChromeOptions
import pandas as pd
import openpyxl
import os
import tkinter as tk
from tkinter import messagebox

def driversetup(url):
    options = ChromeOptions()
    options.add_experimental_option("detach", True)
    
    # 로컬에 저장된 ChromeDriver 경로를 지정
    driver_path = "./chromedriver.exe"  # chromedriver.exe 파일이 스크립트와 동일한 디렉토리에 존재.
    driver = webdriver.Chrome(service=ChromeService(driver_path), options=options)
    driver.get(url)
    time.sleep(3)
    return driver

def start_crawling(keyword, target):
    naver_url = 'https://www.naver.com'
    driver = driversetup(naver_url)

    # 검색창 찾기
    search_box = driver.find_element(By.NAME, 'query')

    # 키워드 입력 및 검색
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.RETURN)

    # 블로그 탭으로 이동
    time.sleep(2)  # 페이지 로딩 대기
    blog_tab = driver.find_element(By.CSS_SELECTOR, 'a[href*="?ssc=tab.blog.all"]')
    blog_tab.click()

    # 최신순 정렬 옵션 선택
    time.sleep(2)  # 페이지 로딩 대기
    latest_option = driver.find_element(By.XPATH, '//a[@role="option" and contains(text(), "최신순")]')
    latest_option.click()

    # 블로그 리스트 크롤링
    time.sleep(2)  # 페이지 로딩 대기

    blog_links = []
    while len(blog_links) < target:
        blog_posts = driver.find_elements(By.CLASS_NAME, 'bx')
        for post in blog_posts:
            if len(blog_links) >= target:
                break
            try:
                if 'user_info' in post.get_attribute('innerHTML'):
                    user_info = post.find_element(By.CLASS_NAME, 'user_info')
                    link = user_info.find_element(By.TAG_NAME, 'a').get_attribute('href')
                    title_link = post.find_element(By.CLASS_NAME, 'title_link').get_attribute('href')
                    if link not in blog_links:
                        blog_links.append((link, title_link))
            except:
                print("user_info 또는 링크를 찾을 수 없습니다.")

        # 스크롤 내리기
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

    # Targetmail 및 Targetlink 생성
    data = []
    for email_link, blog_link in blog_links:
        blog_id = email_link.split('/')[-1]
        target_mail = f"{blog_id}@naver.com"
        data.append({"Targetmail": target_mail, "Targetlink": blog_link})

    # DataFrame 생성
    df = pd.DataFrame(data)

    # 현재 시간 가져오기
    current_time = time.strftime("%Y%m%d_%H%M%S")

    # 임시 엑셀 파일로 저장 (하이퍼링크 추가 전)
    temp_filename = f"temp_blogtarget_{current_time}.xlsx"
    df.to_excel(temp_filename, index=False)

    # openpyxl을 사용하여 하이퍼링크 추가 및 열 너비 조정
    wb = openpyxl.load_workbook(temp_filename)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        email = ws[f"A{row}"].value
        email_link = f"mailto:{email}"
        ws[f"A{row}"].hyperlink = email_link
        ws[f"A{row}"].style = "Hyperlink"
        
        target_link = ws[f"B{row}"].value
        ws[f"B{row}"].hyperlink = target_link
        ws[f"B{row}"].style = "Hyperlink"

    # 열 너비 조정
    ws.column_dimensions['A'].width = 30  # Targetmail 열 너비 설정
    ws.column_dimensions['B'].width = 50  # Targetlink 열 너비 설정

    # 최종 엑셀 파일로 저장
    output_filename = f"blogtarget_{current_time}.xlsx"
    wb.save(output_filename)
    os.remove(temp_filename)

    # 파일 경로 확인
    print(f"엑셀 파일이 현재 디렉토리에 저장되었습니다: {os.path.abspath(output_filename)}")

    # 결과 출력
    for i, row in df.iterrows():
        print(f"Targetmail: {row['Targetmail']}, Targetlink: {row['Targetlink']}")

    # 드라이버 종료
    driver.quit()

def get_user_input():
    user_input = {}

    def on_submit():
        user_input['keyword'] = keyword_entry.get()
        try:
            user_input['target'] = int(target_entry.get())
            root.quit()
        except ValueError:
            messagebox.showerror("Error", "수집할 블로그의 수는 숫자여야 합니다.")
            return

    root = tk.Tk()
    root.title("블로그 크롤링 입력")
    root.geometry("450x200")

    tk.Label(root, text="키워드를 입력하세요:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
    keyword_entry = tk.Entry(root, width=40)
    keyword_entry.grid(row=0, column=1, padx=10, pady=10, sticky='w')

    tk.Label(root, text="수집할 블로그의 수:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
    target_entry = tk.Entry(root, width=40)
    target_entry.grid(row=1, column=1, padx=10, pady=10, sticky='w')

    submit_button = tk.Button(root, text="Submit", command=on_submit)
    submit_button.grid(row=2, column=0, columnspan=2, pady=10)

    root.mainloop()

    return user_input

def main():
    user_input = get_user_input()

    if user_input and 'keyword' in user_input and 'target' in user_input:
        start_crawling(user_input['keyword'], user_input['target'])
    else:
        messagebox.showerror("Error", "모든 입력을 완료해 주세요.")

if __name__ == "__main__":
    main()
